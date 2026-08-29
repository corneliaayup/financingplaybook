from __future__ import annotations

import json
from pathlib import Path

import openpyxl

from .input_types import CaseInput, FieldIssue
from .validate import iter_fields, validate_record, validate_value


class ReaderError(Exception):
    """Raised when an input file cannot be read."""


def read_excel_form(path: str | Path) -> dict[str, object]:
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # openpyxl raises several file-specific errors
        raise ReaderError(f"Could not open workbook: {exc}") from exc
    if "Questionnaire" not in wb.sheetnames:
        raise ReaderError("Workbook has no 'Questionnaire' sheet")
    ws = wb["Questionnaire"]
    header = None
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        first = str(row[0]).strip() if row and row[0] is not None else ""
        if first == "No.":
            header = i
            continue
        if header is not None and row[0] is not None and len(row) >= 4:
            rows.append((str(row[0]).strip(), row[3]))
    if header is None:
        raise ReaderError(
            "Could not find header row (column A = 'No.') in Questionnaire sheet"
        )
    out: dict[str, object] = {}
    for no, val in rows:
        if val is None or str(val).strip() == "":
            continue
        out[no] = val
    return out


def read_json_record(src: str | Path) -> dict[str, object]:
    try:
        text = Path(src).read_text() if isinstance(src, Path) else src
    except OSError as exc:
        raise ReaderError(f"Could not read file: {exc}") from exc
    try:
        data = json.loads(text)
    except json.JSONDecodeError as exc:
        raise ReaderError(f"Not valid JSON: {exc}") from exc
    if not isinstance(data, dict):
        raise ReaderError("JSON record must be an object of slug: value")
    return data


def _resolve(
    raw: dict, questionnaire: dict
) -> tuple[dict[str, object], list[FieldIssue]]:
    resolved: dict[str, object] = {}
    issues: list[FieldIssue] = []
    alias_to_slug = {}
    slug_labels = {}
    for field in iter_fields(questionnaire):
        slug = field["slug"]
        slug_labels[slug] = field.get("label", slug)
        alias = (field.get("aliases") or {}).get("form_column_d")
        if alias is not None:
            alias_to_slug[str(alias).strip()] = slug
    for key, value in raw.items():
        slug = None
        if str(key).strip() in alias_to_slug:
            slug = alias_to_slug[str(key).strip()]
        elif str(key) in slug_labels:
            slug = str(key)
        if slug is None:
            issues.append(
                FieldIssue(
                    str(key), str(key), f"Unknown field or question {key!r}", value
                )
            )
        else:
            resolved[slug] = value
    return resolved, issues


def build_case_input(
    raw: dict, questionnaire: dict, source: str
) -> CaseInput:
    resolved, issues = _resolve(raw, questionnaire)
    issues += list(validate_record(resolved, questionnaire))
    record: dict[str, object] = {}
    context: dict[str, object] = {}
    for field in iter_fields(questionnaire):
        slug = field["slug"]
        if slug not in resolved:
            continue
        if field.get("context_source"):
            if validate_value(field, resolved[slug]):
                continue  # invalid -> excluded
            context[slug] = float(str(resolved[slug]).strip())
        elif validate_value(field, resolved[slug]):
            continue  # invalid -> excluded, engine reports insufficient below
        else:
            if field["type"] == "likert_5":
                record[slug] = int(str(resolved[slug]).strip())
            elif field["type"] == "numeric":
                record[slug] = float(str(resolved[slug]).strip())
            elif field["type"] == "choice":
                record[slug] = str(resolved[slug]).strip()
            # text/date metadata fields are known but not scored; kept out of record
    return CaseInput(
        record=record, context=context, issues=tuple(issues), source=source
    )
